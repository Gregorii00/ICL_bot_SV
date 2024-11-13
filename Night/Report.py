import pandas as pd
import openpyxl
import datetime
import os.path
from openpyxl.writer.excel import save_workbook
from additional import src_files, src_week_report, src_week_report2, report_week_name1, report_week_name2
def week_report(FILE_NAME1, FILE_NAME2, FILE_NAME3, start, end):
    FILE_NAME1 = src_files + FILE_NAME1
    FILE_NAME2 = src_files + FILE_NAME2
    FILE_NAME3 = src_files + FILE_NAME3

    training = ['Обучение 1', 'Обучение 2', 'Обучение 3', 'Обучение ICL']
    if os.path.exists(FILE_NAME1) != True:
        return 0
    date_now = datetime.datetime.now()
    date_start = date_now - datetime.timedelta(days=7)
    date_last = date_now - datetime.timedelta(days=1)
    str_date_start = str(date_start.date().day) + '.' + str(date_start.date().month)
    str_date_last = str(date_last.date().day) + '.' + str(date_last.date().month) + '.' + str(date_last.date().year)
    date_name = str_date_start + '-' + str_date_last


    workbook = openpyxl.load_workbook(src_week_report)
    sheet = workbook["Лист"]
    workbook2 = openpyxl.load_workbook(src_week_report2)
    sheet2 = workbook2["Лист"]
    df = pd.read_excel(FILE_NAME1, sheet_name='Лист')
    df2 = pd.read_excel(FILE_NAME1, sheet_name='Лист2')
    df_aht = pd.read_excel(FILE_NAME2, sheet_name='Лист')
    df_missed = pd.read_excel(FILE_NAME3, sheet_name='Лист')
    wb = openpyxl.load_workbook(FILE_NAME1)
    df = df[['Время', 'Оператор', 'Оценка']]
    df2 = df2[['Время', 'Оценка', 'Оператор', 'Итог']]
    df_aht = df_aht[['Оператор', 'Принятые вх.', 'AHT вх., сек']]
    df_missed = df_missed[['Оператор', 'Пропущенные']]
    start = start.replace("-", "")
    end = end.replace("-", "")
    temp = pd.DatetimeIndex(df['Время'])
    df['Date'] = temp.date
    df['Time'] = temp.time
    del df['Время']
    start = datetime.datetime.strptime(start, "%Y%m%d").date()
    end = datetime.datetime.strptime(end, "%Y%m%d").date()
    dates = str(start.day)+'.'+str(start.month) + '-' + str(end.day)+'.'+str(end.month)
    # df = df[(df['Date'] >= start) & (df['Date'] <= end) & (pd.isnull(df['Оценка'])!=True)]
    df = df[(pd.isnull(df['Оценка'])!=True)]
    ind = 0
    for i in df:
        ind+=1
    print(ind)
    list1 = []
    for under_tems, tems in df.items():
        list1.append(tems)
    j=0
    operators={}
    operators_sum={}
    operators_args = {}
    teacher= []
    summ=0
    negativ_no_recalculation = 0
    negativ_recalculation = 0
    for i in df.index:
        operators_sum[list1[0][i]] = operators_sum.get(list1[0][i], 0) + int(list1[1][i])
        operators[list1[0][i]] = operators.get(list1[0][i], 0) + 1
        if int(list1[1][i]) < 4:
            negativ_no_recalculation+=1
        summ+=1
        if list1[0][i] in training:
            a = []
            a.append(list1[0][i])
            a.append(list1[1][i])
            a.append(list1[2][i])
            a.append(list1[3][i])
            teacher.append(a)
    operators2= operators.copy()
    operators_sum2= operators_sum.copy()
    for i in operators_sum:
        operators_args[i] = operators_sum.get(i, 0)
    for i in operators:
        operators_args[i] = operators_args.get(i, 0)/operators.get(i, 0)
    list2 = []
    for under_tems, tems in df2.items():
        list2.append(tems)
    list3 = []
    for under_tems, tems in df_aht.items():
        list3.append(tems)
    operators_aht = {}
    operators_aht_coll = {}
    for i in df_aht.index:
        operators_aht[list3[0][i]] = operators_aht.get(list3[0][i], 0) + int(list3[2][i])* int(list3[1][i])
        operators_aht_coll[list3[0][i]] = operators_aht_coll.get(list3[0][i], 0) + int(list3[1][i])
    list4 = []
    for under_tems, tems in df_missed.items():
        list4.append(tems)
    operators_missed = {}
    for i in df_missed.index:
       operators_missed[list4[0][i]] = list4[1][i]
    for i in range(0, len(list2[3])):
        negativ_recalculation+=1
        if list2[3][i] == 'Необоснованная':
            operators_sum[list2[2][i]] = operators_sum.get(list2[2][i], 0)-list2[1][i]
            operators[list2[2][i]] = operators.get(list2[2][i], 0)-1
    ws = wb.create_sheet(dates)
    ws.cell(row=1, column=1).value = 'Оператор'
    ws.column_dimensions['A'].width = 38
    ws.cell(row=1, column=2).value = 'Оценка сумм до перерасчета'
    ws.cell(row=1, column=3).value = 'Количество до перерасчета'
    ws.cell(row=1, column=4).value = 'Оценка до перерасчета'
    ws.cell(row=1, column=5).value = 'Если нужен перерасчет'
    ws.cell(row=1, column=6).value = 'Сумма всех оценок'
    ws.cell(row=1, column=7).value = 'колличество оценок'
    ws.cell(row=1, column=8).value = 'Оценка'
    ws.cell(row=1, column=9).value = 'Если нужен перерасчет'
    ws.cell(row=1, column=10).value = 'AHT сумма'
    ws.cell(row=1, column=11).value = 'Количество звонков'
    ws.cell(row=1, column=12).value = 'AHT'
    # ws.cell(row=1, column=14).value = 'Пропущенные'
    ws.cell(row=1, column=15).value = 'Из звонков'
    ws.cell(row=2, column=15).value = negativ_no_recalculation
    ws.cell(row=1, column=16).value = 'Из прослушанных'
    ws.cell(row=2, column=16).value = negativ_recalculation
    sheet.cell(row=1, column=1).value = report_week_name1 + date_name
    sheet2.cell(row=1, column=1).value = report_week_name2 + date_name
    row = 2
    for i in operators:
        ws.cell(row=row, column=1).value = i
        sheet.cell(row=row+1, column=1).value = i
        sheet2.cell(row=row+1, column=1).value = i
        ws.cell(row=row, column=2).value = operators_sum2.get(i, 0)
        ws.cell(row=row, column=3).value = operators2.get(i, 0)
        ws.cell(row=row, column=4).value = operators_args.get(i, 0)
        sheet.cell(row=row + 1, column=3).value = operators_args.get(i, 0)
        ws.cell(row=row, column=6).value = operators_sum.get(i, 0)
        ws.cell(row=row, column=7).value = operators.get(i, 0)
        ws.cell(row=row, column=8).value = operators_sum.get(i, 0) / float(operators.get(i, 0))
        sheet.cell(row=row + 1, column=4).value = operators_sum.get(i, 0) / float(operators.get(i, 0))
        ws.cell(row=row, column=10).value = operators_aht.get(i, 0)
        ws.cell(row=row, column=11).value = operators_aht_coll.get(i, 0)
        ws.cell(row=row, column=12).value = operators_aht.get(i, 0)/operators_aht_coll.get(i, 0)
        sheet.cell(row=row + 1, column=2).value = operators_aht.get(i, 0)/operators_aht_coll.get(i, 0)
        # ws.cell(row=row, column=14).value = operators_missed.get(i, 0)
        # sheet.cell(row=row + 1, column=5).value = operators_missed.get(i, 0)
        row += 1
    teacher2=[]
    if teacher != []:
        for i in training:
            ws = wb.create_sheet(dates + i)
            ws.cell(row=1, column=1).value = 'Дата'
            ws.cell(row=1, column=2).value = 'Час'
            ws.cell(row=1, column=3).value = 'Сумма оценок'
            ws.cell(row=1, column=4).value = 'Количество'
            row = 2
            for j in teacher:
                if j[0]==i:
                    ws.cell(row=row, column=1).value = j[2]
                    ws.cell(row=row, column=2).value = j[3]
                    ws.cell(row=row, column=3).value = j[1]
                    ws.cell(row=row, column=4).value = 1
                    row += 1
                    
    save_workbook(wb, FILE_NAME1)
    workbook.save(src_week_report)
    workbook2.save(src_week_report2)
    print("До перерасчета: ", negativ_no_recalculation, " После перерасчета: ", negativ_recalculation)
    print('Всего оценок: ', summ)
    return summ
