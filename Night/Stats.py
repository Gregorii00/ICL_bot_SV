from additional import src_files, src_files_samples
import openpyxl
from openpyxl.styles import Font, PatternFill
def stats_operators_with_missing(FILE_NAME4, FILE_NAME5, FILE_NAME6, miss = True):
    FILE_NAME1 = src_files + FILE_NAME4
    FILE_NAME2 = src_files + FILE_NAME5
    FILE_NAME3 = src_files + FILE_NAME6
    if miss:
        FILE_STAT_OPER = src_files_samples + 'Статистика операторов с пропущенными.xlsx'
    else:
        FILE_STAT_OPER = src_files_samples + 'Статистика операторов без пропущенных.xlsx'
    workbook = openpyxl.load_workbook(FILE_STAT_OPER)
    sheet = workbook["Лист"]
    workbook1 = openpyxl.load_workbook(FILE_NAME1)
    sheet1 = workbook1["Лист"]
    workbook2 = openpyxl.load_workbook(FILE_NAME2)
    sheet2 = workbook2["Лист"]
    workbook3 = openpyxl.load_workbook(FILE_NAME3)
    sheet3 = workbook3["Лист"]
    stats = []

    for i in range(2, 50):
        if sheet1.cell(row=i, column=1).value:
            k = []
            k.append(sheet1.cell(row=i, column=1).value)
            k.append(sheet1.cell(row=i, column=2).value)
            k.append(sheet1.cell(row=i, column=3).value)
            k.append(sheet1.cell(row=i, column=4).value)
            k.append(sheet1.cell(row=i, column=5).value)
            k.append(sheet1.cell(row=i, column=6).value)
            k.append(sheet1.cell(row=i, column=7).value)
            k.append(sheet1.cell(row=i, column=8).value)
            k.append(sheet1.cell(row=i, column=9).value)
            k.append(sheet1.cell(row=i, column=10).value)
            stats.append(k)

    max_count_id_lider = [0, 0, 0]
    for j in range(3):
        max_count = int(stats[0][1])
        max_count_id = 0
        for i in range(0, len(stats)):
            if j == 0:
                if max_count < int(stats[i][1]):
                    max_count = int(stats[i][1])
                    max_count_id = i
            else:
                if max_count < int(stats[i][1]) and int(stats[i][1]) < int(stats[max_count_id_lider[j-1]][1]):
                    print(stats[max_count_id_lider[j-1]][1])
                    max_count = int(stats[i][1])
                    max_count_id = i
        max_count_id_lider[j] = max_count_id
    for i in range(0, len(stats)):
        for j in range(0, len(stats)):
            if sheet2.cell(row=j+2, column=3).value and sheet2.cell(row=j+2, column=3).value == stats[i][0]:
                stats[i].append(sheet2.cell(row=j+2, column=4).value)
    if miss:
        for i in range(0, len(stats)):
            for j in range(0, len(stats)):
                if sheet3.cell(row=j + 2, column=3).value and sheet3.cell(row=j + 2, column=3).value == stats[i][0]:
                    stats[i].append(sheet3.cell(row=j + 2, column=4).value)
    for i in range(0, len(stats)):
        sheet.cell(row=i+3, column=1).value = stats[i][0]
        sheet.cell(row=i+3, column=2).value = stats[i][1]
        if i == max_count_id_lider[0]:
            sheet.cell(row=i+3, column=2).fill = PatternFill('solid', fgColor="00B04E")
        elif i == max_count_id_lider[1] or i == max_count_id_lider[2]:
            sheet.cell(row=i + 3, column=2).fill = PatternFill('solid', fgColor="92D050")
        else:
            sheet.cell(row=i + 3, column=2).fill = PatternFill('solid', fgColor="FFFFFF")
        sheet.cell(row=i+3, column=3).value = stats[i][2]
        if int(stats[i][2])<=140:
            sheet.cell(row=i+3, column=3).fill = PatternFill('solid', fgColor="00B04E")
        elif int(stats[i][2])>140 and int(stats[i][2])<=200:
            sheet.cell(row=i + 3, column=3).fill = PatternFill('solid', fgColor="92D050")
        elif int(stats[i][2])>200 and int(stats[i][2])<=240:
            sheet.cell(row=i + 3, column=3).fill = PatternFill('solid', fgColor="FFFF00")
        elif int(stats[i][2])>240 and int(stats[i][2])<=340:
            sheet.cell(row=i + 3, column=3).fill = PatternFill('solid', fgColor="FF0000")
        elif int(stats[i][2])>340:
            sheet.cell(row=i + 3, column=3).fill = PatternFill('solid', fgColor="C00000")
        sheet.cell(row=i+3, column=4).value = stats[i][3]
        sheet.cell(row=i+3, column=5).value = stats[i][4]
        sheet.cell(row=i+3, column=6).value = stats[i][5]
        sheet.cell(row=i+3, column=7).value = stats[i][6]
        sheet.cell(row=i+3, column=8).value = stats[i][7]
        sheet.cell(row=i+3, column=9).value = stats[i][8]
        sheet.cell(row=i+3, column=10).value = stats[i][9]
        sheet.cell(row=i+3, column=11).value = stats[i][10]
        if float(stats[i][10])==5.0:
            sheet.cell(row=i + 3, column=11).fill = PatternFill('solid', fgColor="92D050")
        elif float(stats[i][10])<5.0 and float(stats[i][10])>=4.0:
            sheet.cell(row=i + 3, column=11).fill = PatternFill('solid', fgColor="FFFF00")
        elif float(stats[i][10])<4.0:
            sheet.cell(row=i + 3, column=11).fill = PatternFill('solid', fgColor="FF0000")
    workbook.save(FILE_STAT_OPER)